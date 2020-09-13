import openpyxl as px

class Excel:
    def __init__(self, file_name):
        """
        コンストラクタ
        :param file_name:エクセルファイルパス
        """
        self.file_name = file_name

    def get_excel_data(self, sheet_name='Sheet1'):
        """
        エクセルファイルの特定シートから情報を取得する
        :param sheet_name:
        :return: 取得したエクセルデータ
        """
        res = []
        wb = px.load_workbook(self.file_name)
        ws = wb.get_sheet_by_name(sheet_name)

        for i in range(2,1000):
            jan_cell = 'A' + str(i)
            pref_cell = 'B' + str(i)
            type_cell = 'C' + str(i)
            if ws[jan_cell].value != None and ws[pref_cell].value != None:
                res.append({'jan_code': ws[jan_cell].value, 'area': ws[pref_cell].value, 'type': ws[type_cell].value})
        return res

if __name__ == '__main__':
    ex = Excel('../input/Book.xlsx')
    print(ex.get_excel_data('Sheet1'))
